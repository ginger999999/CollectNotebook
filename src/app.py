
# 批次儲存裁切區塊與正確答案

from flask import Flask, render_template, request, redirect, url_for, flash, send_from_directory
import os
from PIL import Image
import pytesseract
import openpyxl
import random

app = Flask(__name__)
app.secret_key = 'your_secret_key'
UPLOAD_FOLDER = 'src/static/uploads'
EXCEL_FILE = 'wrong_questions.xlsx'

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

def create_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['題目', '解析', '標註'])
        wb.save(EXCEL_FILE)

@app.route('/')
def index():
    return render_template('index.html')
@app.route('/questions')
def questions():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    questions = list(ws.iter_rows(min_row=2, values_only=True))
    return render_template('questions.html', questions=questions)

@app.route('/upload', methods=['POST'])
def upload():
    files = request.files.getlist('file')
    if not files or files[0].filename == '':
        flash('未選擇檔案')
        return redirect(url_for('index'))
    allowed_ext = ['.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.mpo']
    saved_files = []
    for file in files:
        ext = os.path.splitext(file.filename)[1].lower()
        if ext not in allowed_ext:
            continue
        filepath = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(filepath)
        saved_files.append(file.filename)
    if not saved_files:
        flash('沒有可用的圖片格式')
        return redirect(url_for('index'))
    if len(saved_files) == 1:
        return redirect(url_for('crop', filename=saved_files[0]))
    else:
        # 多檔案導向 crop_multi
        return render_template('crop_multi.html', files=saved_files)
@app.route('/crop/<filename>')
def crop(filename):
    return render_template('crop.html', filename=filename)

@app.route('/save_crop', methods=['POST'])
def save_crop():
    cropped = request.files['cropped_image']
    origin = request.form.get('origin')
    save_name = f"crop_{origin}_{random.randint(1000,9999)}.png"
    save_path = os.path.join(UPLOAD_FOLDER, save_name)
    cropped.save(save_path)
    # 可在此記錄到 Excel 或資料庫
    return f'裁切圖片已儲存：{save_name}'

@app.route('/save_crops', methods=['POST'])
def save_crops():
    files = request.files.getlist('cropped_images')
    answers = request.form.getlist('answers')
    origin = request.form.get('origin')
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    saved_names = []
    for idx, f in enumerate(files):
        save_name = f"crop_{origin}_{random.randint(1000,9999)}_{idx}.png"
        save_path = os.path.join(UPLOAD_FOLDER, save_name)
        f.save(save_path)
        answer = answers[idx] if idx < len(answers) else ''
        ws.append([save_name, answer, ''])  # 檔名、正確答案、標註
        saved_names.append(save_name)
    wb.save(EXCEL_FILE)
    return f'已儲存 {len(saved_names)} 個裁切區塊到 Excel！'

@app.route('/add', methods=['POST'])
def add():
    question = request.form.get('question')
    solution = request.form.get('solution', '')
    tag = request.form.get('tag', '')
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([question, solution, tag])
    wb.save(EXCEL_FILE)
    flash('已加入Excel錯題本')
    return redirect(url_for('index'))


@app.route('/random', methods=['GET', 'POST'])
def random_question():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    questions = list(ws.iter_rows(min_row=2, values_only=True))
    if not questions:
        return render_template('random.html', questions=None, num=0)
    num = 1
    if request.method == 'POST':
        try:
            num = int(request.form.get('num', 1))
        except ValueError:
            num = 1
        num = max(1, min(num, len(questions)))
    selected = random.sample(questions, min(num, len(questions)))
    return render_template('random.html', questions=selected, num=num)

# 刪除指定列的錯題
@app.route('/delete_question', methods=['POST'])
def delete_question():
    row_index = request.form.get('row_index')
    print(f"收到刪除請求 row_index: {row_index}")
    try:
        row_index = int(row_index)
    except Exception as e:
        flash(f'row_index 解析失敗: {row_index}, 錯誤: {e}')
        return redirect(url_for('questions'))
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    max_row = ws.max_row
    print(f"Excel 最大行: {max_row}")
    # Excel 第一列是標題，row_index+2 才是正確資料列
    target_row = row_index + 2
    print(f"實際要刪除的 Excel 行: {target_row}")
    if target_row <= max_row:
        ws.delete_rows(target_row)
        wb.save(EXCEL_FILE)
        flash(f'已刪除第 {target_row} 行 (row_index={row_index})')
    else:
        flash(f'刪除失敗，索引超出範圍 (row_index={row_index}, target_row={target_row}, max_row={max_row})')
    return redirect(url_for('questions'))


if __name__ == '__main__':
    create_excel()

# 新增儲存選擇狀態路由
@app.route('/update_selected', methods=['POST'])
def update_selected():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    max_row = ws.max_row
    # 取得所有選擇狀態
    for idx in range(max_row - 1):
        selected = request.form.get(f'selected_{idx}')
        answer = request.form.get(f'answer_{idx}')
        # Excel 第一列是標題，資料列從第2行開始
        row = ws[idx + 2]
        # 假設選擇狀態在第4欄，答案在第2欄
        if len(row) > 3:
            row[3].value = '1' if selected else '0'
        if len(row) > 1 and answer is not None:
            row[1].value = answer
    wb.save(EXCEL_FILE)
    flash('已儲存選擇狀態與答案')
    return redirect(url_for('questions'))



# PDF 產生路由（放在所有 @app.route 之後）
from flask import send_file
import io
from PIL import Image

@app.route('/generate_pdf', methods=['POST'])
def generate_pdf():
    data = request.get_json()
    images = data.get('images', [])
    if not images:
        return 'No images selected', 400
    img_list = []
    for img_url in images:
        filename = img_url.split('/')[-1]
        img_path = os.path.join(UPLOAD_FOLDER, filename)
        if os.path.exists(img_path):
            img = Image.open(img_path).convert('RGB')
            img_list.append(img)
    if not img_list:
        return 'No valid images found', 400
    # 設定 A4 尺寸 (2480x3508 px, 300 dpi)，上下邊距各 36px
    a4_width, a4_height = 2480, 3508
    margin = 36
    content_width = a4_width - 2 * margin
    content_height = a4_height - 2 * margin
    # 依序縮放每張圖至 A4 寬度
    resized_imgs = []
    for img in img_list:
        w, h = img.size
        new_h = int(h * content_width / w)
        resized = img.resize((content_width, new_h), Image.LANCZOS)
        resized_imgs.append(resized)
    # 分頁，每頁內容高度 content_height
    pages = []
    current_imgs = []
    current_height = 0
    for img in resized_imgs:
        if current_height + img.height > content_height and current_imgs:
            # 新頁
            pages.append(current_imgs)
            current_imgs = []
            current_height = 0
        current_imgs.append(img)
        current_height += img.height
    if current_imgs:
        pages.append(current_imgs)
    # 建立每頁圖
    page_imgs = []
    answers = data.get('answers', [])
    answer_height = 80 if pages and answers else 0
    for page_idx, imgs in enumerate(pages):
        # 最後一頁底部預留答案空間
        is_last = (page_idx == len(pages)-1)
        extra_space = answer_height if is_last else 0
        total_h = sum(img.height for img in imgs)
        merged_img = Image.new('RGB', (a4_width, a4_height), (255,255,255))
        y_offset = margin + (content_height - total_h - extra_space)//2 if total_h + extra_space < content_height else margin
        for img in imgs:
            x_offset = margin + (content_width - img.width)//2
            merged_img.paste(img, (x_offset, y_offset))
            y_offset += img.height
        # 最後一頁加上答案欄
        if is_last and answers:
            from PIL import ImageDraw, ImageFont
            draw = ImageDraw.Draw(merged_img)
            font_path = os.path.join(os.path.dirname(__file__), 'fonts', 'arial.ttf')
            try:
                font = ImageFont.truetype(font_path, 80)
            except Exception as e:
                font = ImageFont.load_default()
            ans_y = a4_height - margin - answer_height + 10
            ans_text = "Ans: " + ", ".join([f"{idx+1}. {ans}" for idx, ans in enumerate(answers)])
            draw.text((margin, ans_y), ans_text, fill=(0,0,0), font=font)
        page_imgs.append(merged_img)
    # 儲存為多頁 PDF
    pdf_bytes = io.BytesIO()
    page_imgs[0].save(pdf_bytes, format='PDF', save_all=True, append_images=page_imgs[1:])
    pdf_bytes.seek(0)
    return send_file(pdf_bytes, mimetype='application/pdf', as_attachment=False, download_name='marked_images.pdf')

if __name__ == '__main__':
    create_excel()
    app.run(host='0.0.0.0', port=5000, debug=True)
