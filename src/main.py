import os
from PIL import Image
import pytesseract
import openpyxl
import random

EXCEL_FILE = 'wrong_questions.xlsx'

# 建立 Excel 錯題本
def create_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['題目', '解析', '標註'])
        wb.save(EXCEL_FILE)

# 從圖片擷取題目文字
def extract_text_from_image(image_path):
    img = Image.open(image_path)
    text = pytesseract.image_to_string(img, lang='chi_tra')
    return text

# 新增錯題到 Excel
def add_wrong_question(question, solution='', tag=''):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([question, solution, tag])
    wb.save(EXCEL_FILE)

# 隨機出題
def random_question():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    questions = list(ws.iter_rows(min_row=2, values_only=True))
    if not questions:
        print('目前沒有錯題')
        return
    q = random.choice(questions)
    print(f'題目: {q[0]}')
    print(f'解析: {q[1]}')
    print(f'標註: {q[2]}')

# 主選單
if __name__ == '__main__':
    create_excel()
    while True:
        print('\n1. 從圖片擷取錯題並加入Excel')
        print('2. 隨機出題')
        print('3. 離開')
        choice = input('請選擇功能: ')
        if choice == '1':
            img_path = input('請輸入圖片路徑: ')
            text = extract_text_from_image(img_path)
            print('擷取到的文字:')
            print(text)
            confirm = input('是否要加入Excel？(y/n): ')
            if confirm.lower() == 'y':
                add_wrong_question(text)
                print('已加入Excel錯題本')
        elif choice == '2':
            random_question()
        elif choice == '3':
            break
        else:
            print('請輸入正確選項')
